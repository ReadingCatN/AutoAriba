from msilib.schema import Directory
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException,NoSuchElementException,StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
#from selenium.webdriver.common.action_chains import ActionChains
from tkinter import filedialog, messagebox,simpledialog
from openpyxl.styles import PatternFill,Font, Alignment,Side,Border
from openpyxl import load_workbook
import openpyxl,os,time,re,sys,shutil,unicodedata,glob
import pandas as pd 
import tkinter as tk
import win32com.client as win32



ARIBA_URL="http://covestro-child1.procurement-eu.ariba.com/"
ARIBA_ADMIN_DIC={'CVTWX':'cassie.song@covestro.com','CVSBZ':'charlotte.chen@covestro.com','CXGUQ':'michelle.zhang@covestro.com','CVSYX':'peng.gao@covestro.com'}
ARIBA_ADMIN_LIST=['CVTWX','CVSBZ','CXGUQ']
CURRENCY_LISTS=['USD','EUR','TWD','CNY','JPY','HKD','THB']
EXCLUDE_TYPE_LISTS=['Punchout L2','Punchout L1','WIP Work in Progress']
CAT_REGION=['APAC','CN','HK','JP','TH','TW']
STANDARD_COLUMNS = ['Supplier ID', 'Supplier Part ID', 'Item Description', 'Unit Price',
            'Unit of Measure', 'Short Name','Domain', 'Value','Lead Time','Currency' ]

class Ariba_Auto:
    def __init__(self):
        self.ariba_site=ARIBA_URL
        self.cat_status=False
        self.mail_status=False
        self.compare_status=False
        self.user_account=os.getenv('USERNAME')
        self.init_check()

        
    """
    ==========Initial Self-check is Required ============ 

    """
    def base_path_get(self):
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
            self.running_status='exe'
        else:
            base_path = os.path.abspath(".")
            self.running_status='script'
        return base_path

    def init_check(self):
        base_path=self.base_path_get()
        # if self.running_status == 'exe':
        # # When running as an .exe, files are in a temporary directory
        #     asset_path = os.path.join(sys._MEIPASS, "asset")
        # else:
        asset_path = os.path.join(base_path, "asset")
        
        report_path=os.path.join(base_path,"report")

        cat_onedriver_shortcut=rf'C:\Users\{self.user_account}\OneDrive - Covestro\Catalogs\Catalog Tracker.xlsx'
        if os.path.exists(cat_onedriver_shortcut):
            print("Main Cat Data Check")
            print("==========================================")
            self.cat_onedriver_shortcut=cat_onedriver_shortcut
        else:
            messagebox.showerror("File Not Found", "Main Cat Data does not exist.")
            sys.exit("Initialization failed: Required file not found.")
        
        
        edge_drive_path=os.path.join(asset_path, "edgedriver.exe")
        if os.path.exists(edge_drive_path):
            print("Edge Driver File Check")
            print("==========================================")
        elif self.running_status=='exe':
            asset_temp_path=os.path.join(sys._MEIPASS, "asset")
            edge_drive_path=os.path.join(asset_temp_path,"edgedriver.exe")
        elif self.running_status=='script':
            messagebox.showerror("File Not Found", "Edge Driver needs to be downloaded first")
            sys.exit("Initialization failed: edge driver not found.")
        self.driver_path=edge_drive_path
        
        config_list=os.path.join(asset_path, "config.xlsx") 
        if os.path.exists(config_list):
            print("Config List Check")
            print("==========================================")
            self.config_list=config_list
            self.uom_prefer_list()
        elif self.running_status == 'script':
            messagebox.showerror("File Not Found", "You need to build the config list first")
            sys.exit("Initialization failed: config list not found.")
        # this is for further distribution of program for end-users 
        elif self.running_status == 'exe':
            extracted_config_path = os.path.join(base_path, "config.xlsx")
            if not os.path.exists(extracted_config_path):
                asset_temp_path=os.path.join(sys._MEIPASS, "asset")
                config_temp_list=os.path.join(asset_temp_path,"config.xlsx")
                shutil.copy(config_temp_list, extracted_config_path)
                print(f"Extracted config.xlsx to {extracted_config_path}")
            self.config_list = extracted_config_path
            self.uom_prefer_list()

        download_dir = os.path.join(report_path,"Download") 
        if not os.path.exists(download_dir):
            os.makedirs(download_dir)
            print(f"Created folder: {download_dir}")
            print("==========================================")
        self.download_dir=download_dir
        
        download_all_dir =os.path.join(report_path,"AllCatDownload")  
        if not os.path.exists(download_all_dir):
            os.makedirs(download_all_dir)
            print(f"Created folder: {download_all_dir}")
            print("==========================================")
        self.download_all_dir=download_all_dir
        
        report_dir=os.path.join(report_path,"ChangeReport") 
        if not os.path.exists(report_dir):
            os.makedirs(report_dir)
            print(f"Created folder: {report_dir}")
            print("==========================================")
        self.report_dir=report_dir

        merge_dir=os.path.join(report_path,"MergeReport") 
        if not os.path.exists(merge_dir):
            os.makedirs(merge_dir)
            print(f"Created folder: {merge_dir}")
            print("==========================================")
        self.merge_dir=merge_dir

    """
    ======Part 1  Cat quality Check ======
    1.provide 2 modes for choose mode==0 to manually choose the file for quality check 
      mode==1 based on the route in config.xlsx to auto check 
    """
    """
    Get the general cat tracker list
    """
    def cat_tracker_get(self):
        df=pd.read_excel(self.cat_onedriver_shortcut,sheet_name='Main Tracker')
        df=df[df['Country'].isin(CAT_REGION) & ~df['Catalog type'].isin(EXCLUDE_TYPE_LISTS)]
        return df
    """
    Get the UOM requirement (specific)
    """
    def uom_prefer_list(self):
        uom_df=pd.read_excel(self.config_list,sheet_name='uom',skiprows=1)
        self.uom_list=uom_df[uom_df['Preferred']=='Yes']['Value'].tolist()
    """
    mode 0: select file one file per time
    """
    def file_select(self):
        root=tk.Tk()
        root.withdraw()
        file_path=filedialog.askopenfilename(
            title="Please Select the Catalogue",
            filetypes=[("Excel files","*.xlsx *.xls")])
        return file_path
    """
    get the requirements for MG --  just to check if mg in Ariba MG list
    """
    def mg_list_get(self):
        if self.config_list:
            requirement_df=pd.read_excel(self.config_list,sheet_name='mglist')
            mgs_4_check=requirement_df['UniqueName'].to_list()
            return mgs_4_check
            # mgs=self.mgs_4_check
            # input_flag=False
            # while True:
            #     mg=simpledialog.askstring("Input","Please Enter the MGs for Check")
            #     if not mg and input_flag:
            #         break
            #     elif mg and mg.isdigit() and len(mg)==8:
            #         mgs.append(mg)
            #         input_flag=True
            #     else:
            #         messagebox.showerror("Invalid Input", "Please Enter the Right MG")
            # self.mgs_4_check=mgs
        else:
            sys.exit("No Config List", "Please check your config list file")


    """
    Main function for quality check 
    defualt mode==0 initiate the file choose
    mode==1 refer to the config file (excel)
    """
    def static_cat_quality_check(self,mode=0):
        self.uom_prefer_list()
        mg_lists=self.mg_list_get()
        if mode==0:
            file_path=self.file_select()
            
            print(f"In mode {mode}, all the rquirements for quality check have been get")          
            self.loop_excel_check(file_path,mg_lists)
        elif mode==1:
            requirement_df=pd.read_excel(self.config_list,sheet_name='quality')
            print(f"In mode {mode}, all the rquirements for quality check have been get")      
            for _,row in requirement_df.iterrows():
                file_path=row['Route']
                if not self.file_path:
                    messagebox.showerror("File Not Found", "Please input the file route in the config file")
                    return
                self.loop_excel_check(file_path,mg_lists)
        if self.cat_status:
            messagebox.showinfo("Job Done", "You can upload the files") 
        else:
            messagebox.showinfo("Warning", "Modify the files before upload") 
        self.cat_status=False
    """
    Quality check in the excel
    """
    def get_character_count(self,text):
        return sum(1 for char in text if unicodedata.category(char) != 'Cn')

    def loop_excel_check(self,path,mg_lists):
        if path:
            _, filename = os.path.split(path)
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            normal_fill = PatternFill()
            erro_count=0
            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=10):
                supplier_id = row[0].value  # Column A
                part_id=str(row[1].value) #column B 
                uom=row[4].value
                short_name=row[5].value
                class_codes=row[6].value
                mg_no=str(row[7].value)
                currency_value = row[9].value  # Column J
                if 'INT'.lower() in filename.lower():
                    if not (isinstance(supplier_id, str) and supplier_id.isdigit() and len(supplier_id) == 10):
                        row[0].fill = fill
                        erro_count+=1
                    else:
                        row[0].fill=normal_fill
                elif 'EXT'.lower() in filename.lower():
                    if not (isinstance(supplier_id, str) and re.match(r'^AN\d{11}$', supplier_id)):
                        row[0].fill = fill
                        erro_count+=1
                    else:
                        row[0].fill=normal_fill
                if not (len(part_id)<255 and part_id):
                    row[1].fill = fill 
                    erro_count+=1
                else:
                    row[1].fill=normal_fill

                if uom not in self.uom_list:
                    row[4].fill = fill 
                    erro_count+=1
                else:
                    row[4].fill=normal_fill

                if not (self.get_character_count(short_name) <= 40 and short_name):
                    row[5].fill = fill 
                    erro_count+=1
                else:
                    row[5].fill=normal_fill
                
                if not(class_codes=='custom'):
                    row[6].fill = fill 
                    erro_count+=1
                else:
                    row[6].fill=normal_fill
                
                if mg_no not in mg_lists:
                    row[7].fill = fill 
                    erro_count+=1
                else:
                    row[7].fill=normal_fill

                if not(currency_value in CURRENCY_LISTS):
                    row[8].fill = fill 
                    erro_count+=1
                else:
                    row[8].fill=normal_fill

            if erro_count==0:
                self.cat_status=True 
                #messagebox.showinfo("Job Done", "You can upload the file") 
                print(f'This file {path} is ready for upload')
            else:
                print(f'This file {path} needs to be modified before upload')
                #messagebox.showinfo("Warning", "You need to modify the file") 
            
            # directory, filename = os.path.split(path)

            # # Create the new filename
            # new_filename = "processed_" + filename

            # # Join the directory with the new filename to get the full new path
            # new_path = os.path.join(directory, new_filename)

            # Save the workbook with the new path
            workbook.save(path)

            # Print the new path for confirmation
            print("==========================================")
    
    """
    ======Part 2  Ariba page operation ======
    Acted as the Admin and reach the page of core admin for further action
    """

    """
    Log into Ariba Site
    """
    def ariba_admin_login(self,mode=0):
        options = Options()
        if mode==0:
            options.add_experimental_option("prefs", {
                "download.default_directory": self.download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            })
        elif mode==1:
            options.add_experimental_option("prefs", {
                "download.default_directory": self.download_all_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            })
        service = Service(self.driver_path)
        driver = webdriver.Edge(service=service,options=options)
        self.driver=driver
        driver.get(self.ariba_site)
        # need to judge whether a pop up window for different account sign-in
        try:
            # Check if the page title is "Sign in to your account"
            WebDriverWait(self.driver, 5).until(
                EC.title_contains("Sign in to your account")
            )
            print("Sign in page detected.")
            # Perform sign-in actions here
            log_account= ARIBA_ADMIN_DIC[self.user_account]
            email_element = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, f"//small[text()='{log_account}']")))
    # Find the parent element that needs to be clicked
            parent_element = email_element.find_element(By.XPATH, "./ancestor::div[@class='table-row']")
    # Click the parent element
            parent_element.click()
            print(f"Already Sign in with account {log_account}.")
            print("==========================================")
        except TimeoutException:
            print("Sign in page not detected.")
            print("==========================================")

        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            print("Page fully loaded")
            print("==========================================")
        except TimeoutException:
            messagebox.showwarning("Connection Problem", "Please Retry the program")  
        

    """
    Log out of Ariba Site
    """    
    def ariba_admin_logout(self):
        driver=self.driver
        if driver:
            driver.quit()
        print("Page log out")
    
    """
    supporting function for web ope. 

    """
    
    def wait_and_click(self, by, value, timeout=20,type=0):
        try:
            if self.is_element_in_iframe(by,value):
                iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                for iframe in iframes:
                    self.driver.switch_to.frame(iframe)
                    try:
                        self.driver.find_element(by, value)
                        break
                    except NoSuchElementException:
                        self.driver.switch_to.default_content()
                        continue
            element = WebDriverWait(self.driver, timeout).until(
                EC.element_to_be_clickable((by, value))
            )
            element.click()
        except TimeoutException:
            #type==0 for general purpose, will pop up the warning
            if type==0:
                messagebox.showwarning("Connection Problem", f"Element with {by}={value} not found within {timeout} seconds.")
            elif type==1:
                pass
    
    """
    To judge if element is in a iframe
    """
    def is_element_in_iframe(self,by,value, timeout=20):
        try:
            # Check if the element is present in the main document
            self.driver.find_element(by, value)
            return False
        except NoSuchElementException:
            pass

        # Iterate through all iframes and check if the element is present in any of them
        # iframes = WebDriverWait(self.driver, timeout).until(
        #     EC.presence_of_all_elements_located((By.TAG_NAME, "iframe"))
        # )
        iframes=self.driver.find_elements(By.TAG_NAME, "iframe")
        for index, iframe in enumerate(iframes):
            self.driver.switch_to.frame(iframe)
            try:
                self.driver.find_element(by, value)
                print(f"Element found in iframe {index}: {by}={value}")
                return True
            except NoSuchElementException:
                print(f"Element not found in iframe {index}: {by}={value}")
                self.driver.switch_to.default_content()
        return False
    
    """
    supporting function for web ope. 
    
    """
    def input_and_search(self, by, value,input,timeout=20,type=0):
        try:
            if self.is_element_in_iframe(by,value):
                iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                for iframe in iframes:
                    self.driver.switch_to.frame(iframe)
                    # print('Yes, it is a iframe')
                    try:
                        self.driver.find_element(by, value)
                        break
                    except NoSuchElementException:
                        self.driver.switch_to.default_content()
                        continue
            element = WebDriverWait(self.driver, timeout).until(
                EC.visibility_of_element_located((by, value))
            )
            element.clear()
            if type==0:
                element.click()
                element.send_keys(input)
            elif type==1:
                print(input)
                element.send_keys(input)
                #element.send_keys(Keys.RETURN)
        except TimeoutException:
            messagebox.showwarning("Connection Problem", f"Element with {by}={value} not found within {timeout} seconds.")
        except Exception as e:
            print(f"An error occurred: {e}")

    """
    Main function part 1 reach the page first
    """
    def ariba_catpage_get(self,flag=0,timeout=20):
        driver=self.driver
        if self.user_account not in ARIBA_ADMIN_LIST and driver:    
            try:
                delegate_element = WebDriverWait(driver, timeout).until(
                    EC.element_to_be_clickable((By.ID, "_tdagcd"))
                )
                for _ in range(100): 
                    delegate_element.click()
            except TimeoutException:
                sys.exit("Auth. Problem: You are not authorized.")
            except ElementClickInterceptedException:
                driver.execute_script("arguments[0].click();", delegate_element)
            except StaleElementReferenceException:
                pass
        #also need to judge whether account has delegate the account to others
        elif self.user_account in ARIBA_ADMIN_LIST and driver:
            try:
                delegate_element = WebDriverWait(driver, timeout=5).until(
                    EC.element_to_be_clickable((By.XPATH, "//table[@class='w-arw-page-template']//div[@class='pageHead w-page-head' and contains(text(), 'Delegation')]"))
                )
                print('Delagation Page found')
                try:
                    continue_button = WebDriverWait(driver, timeout).until(
                        EC.element_to_be_clickable((By.XPATH, "//a[text()='Continue' and @bh='HL']")))
                    print("Continue button found.")
                    for _ in range(100):
                        continue_button.click()
                except TimeoutException:
                    print('Continue button cant find')
            except TimeoutException:
                print("User didn't delegate the account")
        
        
            #Wait for the button to be present
        try:
            continue_button = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.ID, "_bf7aib")))
        # Click the button
            continue_button.click()
        except TimeoutException:
            print('User dont have OK confirmation')
        WebDriverWait(driver,timeout=5)
        driver.maximize_window()
        if flag==0:
            actions = [
                (By.ID, "_dbw$v"),  # Site link
                (By.ID, "__uxijd"),  # Pop-up element
                (By.ID, "_s2d3v"),  # Manage link
                (By.CSS_SELECTOR, 'a[title="Core Administration"]'),  # Core Administration link
                (By.CSS_SELECTOR, 'a[title="Catalog Manager"]'),# Catalog Manager link
                #(By.ID, "_6xw98d"),  # Catalog Manager link
                #(By.ID, "_xzb03d")   # Catalogs link
                (By.LINK_TEXT, "Catalogs") # Catalogs link
            ]
            for by,value in actions:
                self.wait_and_click(by,value)
        elif flag==1:
            actions = [
                (By.ID, "_dbw$v"),  # Site link
                (By.ID, "__uxijd"),  # Pop-up element
                (By.ID, "_s2d3v"),  # Manage link
                (By.CSS_SELECTOR, 'a[title="Core Administration"]'),  # Core Administration link
                (By.CSS_SELECTOR, 'a[title="Catalog Content Manager"]'),# Catalog content Manager link
                #(By.ID, "_rrfnsb"),  # Catalog content Manager link
                (By.LINK_TEXT, "Content Documents") # Catalogs link
                #(By.ID, "_apbzq")   # Content Documents link
            ]

            for by,value in actions:
                self.wait_and_click(by,value)
            
            iframe_locator = (By.TAG_NAME, "iframe")
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located(iframe_locator))

            # Check if the iframe element is loaded
            iframe_loaded = self.is_element_in_iframe(By.ID, "content-grid")
            if iframe_loaded:
                print("Iframe element is loaded and ready for further actions.")
            else:
                print("Iframe element is not loaded.")
            
        print("Already reach the page, waiting for further action")
        print("==========================================")
    
    """
    This part needs to be check with Cassie

    """
    
    def ariba_cat_upload(self,mode=0):
        driver=self.driver
        if mode==0:
            cat_lists=[]
            upload_lists=[]
            while True:
                cat_list=simpledialog.askstring("Input","Please Enter the Catalog Name for Search")
                root=tk.Tk()
                root.withdraw()
                upload_list=filedialog.askopenfilename(title="Please Select the Catalogue",filetypes=[("Excel files","*.xlsx *.xls")])
                if not cat_list and not upload_list and len(cat_lists)>=1 and len(upload_lists)>=1:
                    break
                elif not cat_list or not upload_list:
                    messagebox.showwarning("Warning", "In current mode, you need to input the cat and choose the file for uploading")
                else:
                    cat_lists.append(cat_list)
                    upload_lists.append(upload_list)

        elif mode==1:
            requirement_df=pd.read_excel(self.config_list,sheet_name='quality')  
            cat_lists=requirement_df['Catalog Subscription Name '].to_list()
            upload_lists=requirement_df['Route'].to_list()
            if not cat_lists or not upload_lists:
                messagebox.showwarning("Warning", "Please check your config document")
                return
        error_log_path=os.path.join(self.download_dir, 'error_config.txt')
        with open(error_log_path, 'w', encoding='utf-8') as log_file:
            pass
        count=0
        for cat_name,file_path in zip(cat_lists,upload_lists):
            self.input_and_search(By.ID,"_yxophd",cat_name)
            self.wait_and_click(By.ID, "_xvmt6c")
            try:
                WebDriverWait(driver, timeout=5).until(
                    EC.presence_of_element_located((By.XPATH, '//td[@class="empty tableBody w-tbl-empty" and text()="No items"]'))
                )
                error_message=(f"{cat_name} not found")
                print(error_message.strip())
                with open(error_log_path,'a', encoding='utf-8') as log_file:
                    log_file.write(error_message +'\n')
                continue
                
            except TimeoutException:
                try:
                    table = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "_nb8ucc")))
                    first_version_link = table.find_element(By.XPATH, ".//a[contains(text(), 'Version')]") 
                    first_version_link.click()
                except NoSuchElementException:
                    continue
                # rows = table.find_elements(By.XPATH, './/tr[contains(@class, "tableRow1")]')
                # for row in rows:
                #     try:
                #         status = row.find_element(By.XPATH, './/td[@class="tableBody w-tbl-cell"]//a[@id="_lfned"]').text
                #         # if status == "Activated":
                #         version_link = row.find_element(By.XPATH, './/td[@class="tableBody w-tbl-cell"]//a[@id="_ybfsed"]')
                #         version_link.click()
                #         # print("Clicked the version link for the first activated row")
                #         break
                #     except NoSuchElementException:
                #         continue
                
                #***according to Cassie, no need to check the quality again here***
                #self.static_cat_quality_check(mode=mode)
                #if self.cat_status:
                
                try:
                    import_tool_element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//span[@title='Import tool']")))
                    import_tool_element.click()
                    file_input_element = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, "//span[@class='w-file-upload']//input[@type='file']"))
                    )
                    file_input_element.send_keys(file_path)
                    count=count+1

                except TimeoutException:
                    pass
                print(f"File {cat_name} already uploaded")

                # catalog_home_button = WebDriverWait(driver, 20).until(
                # EC.presence_of_element_located((By.XPATH, "//button[span[text()='Catalog Home']]")))
                catalog_home_button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//button[@title='Return to the catalog main screen']")))
                catalog_home_button.click()
                # *********to click the upload 
                # import_button = WebDriverWait(driver, 20).until(
                # EC.presence_of_element_located((By.XPATH, "//button[@title='Import new data']")))
                # actions = ActionChains(driver)
                # actions.move_to_element(import_button).perform()
                # import_button = WebDriverWait(driver, 20).until(
                #     EC.element_to_be_clickable((By.XPATH, "//button[@title='Import new data']")))
                # import_button.click()
        
        print("==========================================")
        messagebox.showinfo("Uploaded Success",f"{count} files have been uploaded")
    
    """
    Downloading file function and support function 

    """
    
    def wait_for_download(self, directory, timeout=60):
        seconds = 0
        dl_wait = True
        while dl_wait and seconds < timeout:
            time.sleep(1)
            dl_wait = False
            for fname in os.listdir(directory):
                if fname.endswith('.crdownload'):  # Check for incomplete download file
                    dl_wait = True
            seconds += 1
        return not dl_wait
    
    def action_download(self,row,mode=0,timeout=20):
        # if str_status=='Activated':
        #     status = row.find_element(By.XPATH, ".//div[@col-id='status_label']").text
        #     if status == str_status:
        action_link = row.find_element(By.XPATH, ".//a[@class='actionLink']")
        action_link.click()
        download_original_link = WebDriverWait(self.driver, timeout).until(
        EC.element_to_be_clickable((By.XPATH, "//span[text()='Download Original']/ancestor::a")))
        download_original_link.click()
        if mode==0:
            if self.wait_for_download(self.download_dir):
                print("Download completed successfully.")
                
            else:
                #self.download_status=False
                raise Exception("Download timed out.")
        elif mode==1:
            if self.wait_for_download(self.download_all_dir):
                print("Download completed successfully.")
                
            else:
                #self.download_status=False
                raise Exception("Download timed out.")
            

  
    """
    mode==0  change report 
    mode==1  catlogue download --- for charlotte 
    type==0   with msgbox to input
    type==1 read the config file
    """  

    def init_table_get(self,by1,value1,by2,value2,timeout=20):
        try:
            # WebDriverWait(self.driver, timeout).until(
            # EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
            if self.is_element_in_iframe(by1,value1):
                iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                for iframe in iframes:
                    self.driver.switch_to.frame(iframe)
                    print('Yes, it is a iframe')
                    try:
                        self.driver.find_element(by1, value1)
                        break
                    except NoSuchElementException:
                        self.driver.switch_to.default_content()
                        continue
            ag_grid = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((by1, value1)))
            cat_table = ag_grid.find_element(by2, value2)
            initial_height = cat_table.get_attribute("style").split("height: ")[1].split("px")[0]
            self.initial_height=initial_height
            return cat_table
        except TimeoutException:
            messagebox.showwarning("Connection Problem", f"Element with {by1}={value1} not found within {timeout} seconds.")
        except Exception as e:
            print(f"An error occurred: {e}")

    def get_catalog_name(self):
        try:
            # Adjust the CSS selector to target the correct element
            catalog_name_element = self.driver.find_element(By.CSS_SELECTOR, 'div.ag-cell[col-id="name"]')
            return catalog_name_element.text.strip()
        except:
            return None

    def ariba_cat_download(self,mode=0,input_type=0,timeout=10):
        driver=self.driver
        cat_lists=[]
        if input_type==0:
            input_flag=False
            while True:
                cat_list=simpledialog.askstring("Input","Please Enter the Catalog Name for Search")
                if not cat_list and input_flag:
                    break
                else:
                    cat_lists.append(cat_list)
                    input_flag=True
        elif input_type==1:
            if mode==0:
                requirement_df=pd.read_excel(self.config_list,sheet_name='quality')
                cat_lists=requirement_df['Catalog Subscription Name '].to_list()
                if not cat_lists:
                    messagebox.showwarning("NO CAT", "Please check your config document")
                    return
            elif mode==1:
                cat_df=self.cat_tracker_get()
                cat_lists=cat_df['Catalog Subscription Name '].to_list()
                cat_lists=[cat.strip() for cat in cat_lists if cat]
                if not cat_lists:
                    messagebox.showwarning("NO CAT", "Please check the cat document in SharePoint")
                    return
                
        print(f"Total Cat {len(cat_lists)}")
        error_log_path=os.path.join(self.download_all_dir, 'error_log.txt')
        with open(error_log_path, 'w', encoding='utf-8') as log_file:
            pass  # This will clear the file

        for cat_name in cat_lists:

            try:
            # Wait for the table to be present
                cat_table =self.init_table_get(By.ID,"content-grid",By.CLASS_NAME, "ag-center-cols-container")
                self.input_and_search(By.ID,"searchBox",cat_name,type=1)
                self.wait_and_click(By.XPATH,"//button[@type='submit']")
                #wait for table to change
                #WebDriverWait(driver, timeout).until(self.catalog_name_contains_input(cat_name,driver=driver))
                #WebDriverWait(driver, timeout).until(lambda driver: self.catalog_name_contains_input(cat_name))
                WebDriverWait(self.driver, timeout=5).until(lambda driver: self.get_catalog_name() 
                is not None and cat_name.lower().replace("_", " ") in self.get_catalog_name().lower().replace("_", " "))

            # Find the latest version and latest active version
                rows = cat_table.find_elements(By.XPATH, ".//div[@role='row']")
                latest_version_row = None
                latest_activated_version_row = None
                latest_version = -1
                latest_activated_version=-1
                for row in rows:
                    version_label = row.find_element(By.XPATH, ".//div[@col-id='versionLabel']").text
                    version_number = int(version_label.split()[-1])  # Extract the version number
                    status = row.find_element(By.XPATH, ".//div[@col-id='status_label']").text
                    if version_number > latest_version:
                        latest_version = version_number
                        latest_version_row = row
                    if status == "Activated" and version_number > latest_activated_version:
                        latest_activated_version = version_number
                        latest_activated_version_row = row
                if mode==0:
                    if latest_activated_version_row:
                        self.action_download(latest_activated_version_row)
                    if  latest_version_row and latest_version_row != latest_activated_version_row:
                        self.action_download(latest_version_row)
                    if latest_version_row and latest_activated_version_row and latest_version_row != latest_activated_version_row:
                        self.compare_status=True
                        attached_file=self.ariba_cat_compare()
                        self.email_send(cat_name,attached_file,mode=1)
                    else:
                        print('No need to compare')
                        print("==========================================")

                elif mode==1:
                    if latest_activated_version_row:
                        self.action_download(latest_activated_version_row,mode=mode)

 
            except TimeoutException:
                error_message=(f"{cat_name} not found within {timeout} seconds")
                print(error_message.strip())
                with open(error_log_path,'a', encoding='utf-8') as log_file:
                    log_file.write(error_message +'\n')
            except Exception as e:
                print(f"An error occurred: {e}")
        if mode==0:
            if self.mail_status:
                messagebox.showinfo("Change Report","Reports Generated And Sent")
            else:
                messagebox.showinfo("Change Report","Change Report No Need")
            self.mail_status=False
        elif mode==1:
            self.cat_merge(mode=mode)
            messagebox.showinfo("Merge Report","Merge Reports Generated")
        
        self.mail_status=False
        self.compare_status=False
    
    def cat_merge(self,mode=1):
        if mode==1: 
            merged_file_path = os.path.join(self.merge_dir, 'merged_catalog.xlsx')
            
            # List all .xlsx files in the download directory
            xlsx_files = [f for f in os.listdir(self.download_all_dir) if f.endswith('.xlsx')]
            
            # Initialize an empty DataFrame
            merged_df = pd.DataFrame(columns=STANDARD_COLUMNS)
            
            # Read and concatenate all .xlsx files
            for file in xlsx_files:
                file_path = os.path.join(self.download_all_dir, file)
                print(file)
                df = pd.read_excel(file_path,sheet_name='Items',header=[0,1])
                df=self.cat_data_head_process(df)
                file_name_parts = file.split('_')
                cat_name = '_'.join(file_name_parts[1:-1])
                version = file_name_parts[-1].split('.')[0]
                df['Cat Name'] = cat_name
                df['Version'] = version
                df = df.reindex(columns=STANDARD_COLUMNS+['Cat Name','Version'])
                merged_df = pd.concat([merged_df, df], ignore_index=True)
                
            
            # Save the merged DataFrame to a new .xlsx file
            merged_df.to_excel(merged_file_path, index=False)
            print(f"Merged file saved to {merged_file_path}")
            print("==========================================")
            
            # Delete the original .xlsx files
            for file in xlsx_files:
                file_path = os.path.join(self.download_all_dir, file)
                os.remove(file_path)
                print(f"Deleted file: {file_path}")
        else:
            pass        


    def get_downloaded_files(self, directory,count=2):
        files = [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.xlsx')]
        if len(files)<count:
            return None
        latest_files = sorted(files, key=os.path.getctime, reverse=True)[:count]
        return latest_files


    def cat_data_head_process(self,df):
        df=df.iloc[:,:10]
        new_columns = []
        for col in df.columns:
            
            if col[1] and col[1].strip() and not col[1].startswith('Unnamed'):  # If the second row value exists and is not 'Unnamed'
                new_columns.append(col[1])

            else:  # Otherwise, use the first row value
                new_columns.append(col[0])
        new_columns = self.make_unique(new_columns)
        df.columns = new_columns
        
        return df 
    
    def make_unique(self, columns):
        seen = set()
        result = []
        for col in columns:
            new_col = col
            count = 1
            while new_col in seen:
                new_col = f"{col}_{count}"
                count += 1
            seen.add(new_col)
            result.append(new_col)
        return result

    def ariba_cat_compare(self):
        if self.compare_status:
            file=self.get_downloaded_files(self.download_dir)
            if file and len(file) == 2:
                # Read the downloaded Excel files with Pandas
                df_latest_activated = pd.read_excel(file[0],sheet_name='Items',header=[0,1])
                df_latest = pd.read_excel(file[1],sheet_name='Items',header=[0,1])
                df_latest_activated_process=self.cat_data_head_process(df_latest_activated)
                df_latest_process=self.cat_data_head_process(df_latest)
                # print(df_latest_activated_process.head())
                # print(df_latest_process.head())
                #Get the version of files
                base_version=self.extract_version(file[0])
                new_version=self.extract_version(file[1])
                base_name = os.path.basename(file[0])
                base_name=re.sub(r'_v\d+','',base_name)
                change_report_name=f"Change_Report_V{base_version}&V{new_version}_{base_name}"
                change_report_path = os.path.join(self.report_dir, change_report_name)
                merged_df = pd.merge(df_latest_activated_process, df_latest_process, on='Supplier Part ID', how='outer',
                suffixes=('_old', '_new'),indicator=True)
                merged_df['What Changed'] = ''
                # Identify newly added records
                merged_df.loc[merged_df['_merge'] == 'right_only', 'What Changed'] = 'Added Part'

                # Identify deleted records
                merged_df.loc[merged_df['_merge'] == 'left_only', 'What Changed'] = 'Deleted Part'

                for column in df_latest_activated_process.columns:
                    if column != 'Supplier Part ID':
                        old_col = f"{column}_old"
                        new_col = f"{column}_new"
                        change_col = f"Change on {column}"
                        
                        # Check for changes in each column
                        merged_df[change_col] = merged_df.apply(
                            lambda row: f"Changed from {row[old_col]} to {row[new_col]}" if row[old_col] != row[new_col]  else '', axis=1
                        )
                        
                        # Update 'Change Type' column for modified records
                        merged_df['What Changed'] = merged_df.apply(lambda row: row['What Changed'] if row['What Changed'] in ['Added Part', 'Deleted Part'] 
                        else row['What Changed'] + f" | Change on {column}" if row[change_col] != '' else row['What Changed'], axis=1)
                
                change_report_df = merged_df[['Supplier ID_old', 'Supplier Part ID', 'Item Description_old', 'Short Name_old', 
                              'Unit Price_old', 'Unit Price_new', 'Value_old', 'Value_new', 'Lead Time_old','Lead Time_new','What Changed']]
                # Rename the columns for clarity
                change_report_df.columns = ['Supplier ID', 'Supplier Part ID', 'Item Description', 'Short Name', 
                                            'Old Unit Price', 'New Unit Price', 'Old Value', 'New Value','Old Lead Time','New Lead Time', 'What Changed']
                change_report_df=change_report_df[change_report_df['What Changed']!='']
                change_report_df.to_excel(change_report_path,index=False)
                self.set_report_format(change_report_path)
                print("==========================================")
                return change_report_path


    def extract_version(self, filename):
        match = re.search(r'_v(\d+)', filename)
        return match.group(1) if match else None

    def set_report_format(self,file_path):
    # Load the workbook and select the active worksheet
        wb = load_workbook(file_path)
        ws = wb.active

        # Set the header font to Arial 12 bold
        header_font = Font(name='Arial', size=12, bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # Set the other cells font to Arial 10
        cell_font = Font(name='Arial', size=10)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font

        # Set alignment for all cells
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Auto-fit columns and rows
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    cell_height = len(str(cell.value).split('\n'))
                    if cell_height > max_height:
                        max_height = cell_height
            ws.row_dimensions[row[0].row].height = max_height * 15

        # Save the formatted workbook
        wb.save(file_path)

        print(f"Formatted Excel file saved to {file_path}")
    
    def get_latest_files(self,dir,n=2):
        list_of_files=glob.glob(os.path.join(Directory,'*.xlsx'))
        latest_files=sorted(list_of_files,key=os.path.getatime,reverse=True)[:n]
        return latest_files 


    def email_send(self,cat_name,attachment_path,mode=1):
        if mode==1:
            cat_df=self.cat_tracker_get()
            mail_df=pd.read_excel(self.config_list,sheet_name='email')
            cat_df=cat_df.merge(mail_df,on="Catalog owner",how="left")
            email_add=cat_df[cat_df['Catalog Subscription Name ']==cat_name]['Email'].iloc[0]

            outlook = win32.Dispatch('outlook.application')
    
        # Create a new email
            mail = outlook.CreateItem(0)
            mail.To = email_add
            mail.Subject = f"Change Report of {cat_name}"
            mail.HTMLBody="""<html>
                    <body>
                    <p>Dear Sourcer,</p>
                    <p>Please kindly help to check and confirm newly updated catalog and change report.</p>
                    <p>Thank you!</p>
                    <p>From POA CM Team</p>
                    </body>
                    </html>"""
            
            # Attach the file
            mail.Attachments.Add(attachment_path)
            #attach 2 more catlogue files
            latest_files=self.get_latest_files(self.download_dir,2)
            for file in latest_files:
                mail.Attachments.Add(file)
            mail.Display()
            # Send the email
            #mail.Send()
            print(f"Email sent to {email_add}")
        self.mail_status=True

